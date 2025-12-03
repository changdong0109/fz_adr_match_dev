# -*- coding: utf-8 -*-
"""
匹配结果导出器

功能：
- 导出匹配结果为 Excel（带颜色区分）或 CSV
- 源表字段蓝色，匹配字段黄色高亮，目标表字段绿色
- 支持选择性导出（精确/高置信度/需确认/未匹配）
- 支持合并或分开导出

参考 Step3 的 RelationExporter 实现
"""
import os
import pandas as pd
from typing import List, Dict, Optional, Callable, Tuple


class MatchResultExporter:
    """匹配结果导出器"""
    
    def __init__(self, log_callback: Optional[Callable[[str, str], None]] = None):
        self._log = log_callback or (lambda m, l: None)
    
    def export_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        source_file: str = "",
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> Dict:
        """
        导出为带颜色的 Excel 文件
        
        Args:
            df: 要导出的 DataFrame（列名带 [源]/【匹配】/[目标] 前缀）
            output_path: 输出路径
            source_file: 源表文件名（用于图例）
            progress_callback: 进度回调 (percent, message)
            
        Returns:
            {'success': bool, 'row_count': int, 'message': str}
        """
        progress = progress_callback or (lambda p, m: None)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self._log("openpyxl 未安装，使用普通 Excel 导出", "warning")
            df.to_excel(output_path, index=False)
            return {'success': True, 'row_count': len(df), 'message': f'导出 {len(df)} 条（无颜色）'}
        
        try:
            progress(5, "创建工作簿...")
            wb = Workbook()
            ws = wb.active
            ws.title = "匹配结果"
            
            # 定义样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 蓝色系 - 源表
            header_source_fill = PatternFill(patternType="solid", fgColor="FF93C5FD")
            header_source_font = Font(bold=True, color="FF1E3A8A")
            cell_source_fill = PatternFill(patternType="solid", fgColor="FFDBEAFE")
            
            # 黄色系 - 匹配字段（高亮）
            header_match_fill = PatternFill(patternType="solid", fgColor="FFFCD34D")
            header_match_font = Font(bold=True, color="FF92400E")
            cell_match_fill = PatternFill(patternType="solid", fgColor="FFFEF3C7")
            
            # 绿色系 - 目标表
            header_target_fill = PatternFill(patternType="solid", fgColor="FF86EFAC")
            header_target_font = Font(bold=True, color="FF065F46")
            cell_target_fill = PatternFill(patternType="solid", fgColor="FFD1FAE5")
            
            # 分类列：源表 | 匹配依据字段 | 匹配元信息 | 目标表
            source_cols = []
            match_source_cols = []  # 【匹配源:表名】字段
            match_target_cols = []  # 【匹配目标:表名】字段
            match_meta_cols = []    # 匹配分数、类型、层级
            target_cols = []
            
            for col in df.columns:
                if col.startswith('[源:'):
                    source_cols.append(col)
                elif col.startswith('【匹配源:'):
                    match_source_cols.append(col)
                elif col.startswith('【匹配目标:'):
                    match_target_cols.append(col)
                elif col in ['匹配分数', '匹配类型', '匹配层级']:
                    match_meta_cols.append(col)
                elif col.startswith('[目标:') or col in ['目标表文件', '目标表行号']:
                    target_cols.append(col)
            
            # 确定每列的类型
            col_types = {}
            for col in df.columns:
                if col in source_cols:
                    col_types[col] = 'source'
                elif col in match_source_cols or col in match_target_cols:
                    col_types[col] = 'match'  # 匹配依据字段（黄色高亮）
                elif col in match_meta_cols:
                    col_types[col] = 'match'  # 匹配元信息（黄色）
                elif col in target_cols:
                    col_types[col] = 'target'
                else:
                    col_types[col] = 'match'  # 默认放匹配区
            
            progress(10, "写入表头...")
            
            # 写入表头
            for col_idx, col_name in enumerate(df.columns):
                cell = ws.cell(row=1, column=col_idx + 1, value=col_name)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                col_type = col_types.get(col_name, 'match')
                if col_type == 'source':
                    cell.fill = header_source_fill
                    cell.font = header_source_font
                elif col_type == 'match':
                    cell.fill = header_match_fill
                    cell.font = header_match_font
                else:
                    cell.fill = header_target_fill
                    cell.font = header_target_font
            
            # 写入数据
            total_rows = len(df)
            progress(15, f"写入数据 (共 {total_rows} 行)...")
            
            for row_idx, row in enumerate(df.values, start=2):
                # 更新进度
                if row_idx % 100 == 0:
                    pct = 15 + int((row_idx / total_rows) * 70)
                    progress(pct, f"写入数据 {row_idx}/{total_rows}...")
                
                for col_idx, value in enumerate(row):
                    # 处理 NaN
                    if value is None or (isinstance(value, float) and str(value) == 'nan'):
                        value = ''
                    
                    cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                    cell.border = thin_border
                    
                    col_name = df.columns[col_idx]
                    col_type = col_types.get(col_name, 'match')
                    if col_type == 'source':
                        cell.fill = cell_source_fill
                    elif col_type == 'match':
                        cell.fill = cell_match_fill
                    else:
                        cell.fill = cell_target_fill
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            progress(88, "调整列宽...")
            
            # 自动调整列宽
            for col_idx, col_name in enumerate(df.columns):
                max_len = max(
                    len(str(col_name)),
                    df.iloc[:, col_idx].astype(str).str.len().max() if len(df) > 0 else 0
                )
                ws.column_dimensions[ws.cell(row=1, column=col_idx + 1).column_letter].width = min(max_len + 2, 50)
            
            progress(92, "添加图例说明...")
            
            # 添加图例说明 sheet
            self._add_legend_sheet(wb, source_file)
            
            progress(95, "保存文件...")
            wb.save(output_path)
            
            progress(100, "导出完成")
            
            return {
                'success': True,
                'row_count': len(df),
                'output_path': output_path,
                'message': f'成功导出 {len(df)} 条记录'
            }
            
        except Exception as e:
            self._log(f"Excel导出失败: {e}", "error")
            return {
                'success': False,
                'row_count': 0,
                'message': f'导出失败: {str(e)}'
            }
    
    def _add_legend_sheet(self, wb, source_file: str):
        """添加图例说明 Sheet"""
        from openpyxl.styles import Font, PatternFill
        
        legend_ws = wb.create_sheet("图例说明")
        
        # 样式定义
        source_fill = PatternFill(patternType="solid", fgColor="FF93C5FD")
        source_font = Font(bold=True, color="FF1E3A8A")
        match_fill = PatternFill(patternType="solid", fgColor="FFFCD34D")
        match_font = Font(bold=True, color="FF92400E")
        target_fill = PatternFill(patternType="solid", fgColor="FF86EFAC")
        target_font = Font(bold=True, color="FF065F46")
        
        legend_ws['A1'] = "匹配结果说明"
        legend_ws['A1'].font = Font(bold=True, size=14)
        
        legend_ws['A3'] = "🔵 蓝色区域"
        legend_ws['A3'].fill = source_fill
        legend_ws['A3'].font = source_font
        legend_ws['B3'] = f"源表 ({source_file}) 的字段"
        
        legend_ws['A4'] = "🟡 黄色区域"
        legend_ws['A4'].fill = match_fill
        legend_ws['A4'].font = match_font
        legend_ws['B4'] = "匹配依据字段（源表和目标表用于匹配的字段及匹配值）"
        
        legend_ws['A5'] = "🟢 绿色区域"
        legend_ws['A5'].fill = target_fill
        legend_ws['A5'].font = target_font
        legend_ws['B5'] = "目标表（匹配到的数据）的字段"
        
        legend_ws['A7'] = "列名说明"
        legend_ws['A7'].font = Font(bold=True, size=12)
        
        legend_ws['A8'] = "[源:表名]xxx"
        legend_ws['B8'] = "源表的 xxx 字段（表名在列名中）"
        
        legend_ws['A9'] = "【匹配源:表名】xxx"
        legend_ws['B9'] = "源表用于匹配的字段及匹配值（表名在列名中）"
        
        legend_ws['A10'] = "【匹配目标:表名】xxx"
        legend_ws['B10'] = "目标表用于匹配的字段及匹配值（表名在列名中）"
        
        legend_ws['A11'] = "匹配分数/类型/层级"
        legend_ws['B11'] = "匹配的元信息（分数、类型、层级）"
        
        legend_ws['A12'] = "[目标:表名]xxx"
        legend_ws['B12'] = "目标表的 xxx 字段（表名在列名中）"
        
        legend_ws['A14'] = "匹配层级说明"
        legend_ws['A14'].font = Font(bold=True, size=12)
        
        legend_ws['A15'] = "🟢 精确匹配"
        legend_ws['B15'] = "核心字段完全相等 / POI结构化完全相等，无需人工确认"
        
        legend_ws['A16'] = "🔵 高置信度"
        legend_ws['B16'] = "区县约束 + 模糊匹配 ≥95%，无需人工确认"
        
        legend_ws['A17'] = "🟡 需人工确认"
        legend_ws['B17'] = "模糊匹配 88-95%，建议人工核实"
        
        legend_ws['A18'] = "⚪ 未匹配"
        legend_ws['B18'] = "未找到匹配结果"
        
        # 调整列宽
        legend_ws.column_dimensions['A'].width = 20
        legend_ws.column_dimensions['B'].width = 60
    
    def export_to_csv(
        self,
        df: pd.DataFrame,
        output_path: str,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> Dict:
        """
        导出为 CSV 文件
        
        Args:
            df: 要导出的 DataFrame
            output_path: 输出路径
            progress_callback: 进度回调 (percent, message)
            
        Returns:
            {'success': bool, 'row_count': int, 'message': str}
        """
        progress = progress_callback or (lambda p, m: None)
        
        try:
            progress(10, "准备导出...")
            progress(50, f"写入 {len(df)} 条记录...")
            
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            progress(100, "导出完成")
            
            return {
                'success': True,
                'row_count': len(df),
                'output_path': output_path,
                'message': f'成功导出 {len(df)} 条记录'
            }
        except Exception as e:
            self._log(f"CSV导出失败: {e}", "error")
            return {
                'success': False,
                'row_count': 0,
                'message': f'导出失败: {str(e)}'
            }
    
    def merge_results(
        self,
        result_files: List[str],
        levels: List[str] = None
    ) -> pd.DataFrame:
        """
        合并多个分层结果文件
        
        Args:
            result_files: 结果文件路径列表
            levels: 要合并的层级列表，None 表示全部
            
        Returns:
            合并后的 DataFrame
        """
        level_names = ['精确匹配', '高置信度', '需人工确认', '未匹配']
        if levels:
            level_names = [l for l in level_names if l in levels]
        
        dfs = []
        for f in result_files:
            if not os.path.exists(f):
                continue
            # 检查文件名是否包含需要的层级
            basename = os.path.basename(f)
            for level in level_names:
                if level in basename:
                    try:
                        df = pd.read_csv(f, encoding='utf-8-sig')
                        dfs.append(df)
                    except Exception as e:
                        self._log(f"读取文件失败: {f}, {e}", "warning")
                    break
        
        if dfs:
            return pd.concat(dfs, ignore_index=True)
        return pd.DataFrame()

